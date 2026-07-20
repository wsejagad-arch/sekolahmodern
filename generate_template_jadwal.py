import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Jadwal Guru"

# Header
headers = ["no_induk", "nama_guru", "nama_mapel", "kelas", "hari", "jam_mulai", "jam_selesai", "ruang"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="059669") # Green theme for guru
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
# Column widths
ws.column_dimensions['A'].width = 20 # no_induk
ws.column_dimensions['B'].width = 35 # nama_guru
ws.column_dimensions['C'].width = 20 # nama_mapel
ws.column_dimensions['D'].width = 15 # kelas
ws.column_dimensions['E'].width = 15 # hari
ws.column_dimensions['F'].width = 15 # jam_mulai
ws.column_dimensions['G'].width = 15 # jam_selesai
ws.column_dimensions['H'].width = 20 # ruang

# Borders for headers and 20 empty rows
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=1, max_row=21, min_col=1, max_col=8):
    for cell in row:
        cell.border = thin_border

# Add some sample data to make it clear
sample_data = [
    ["198001012005011001", "KISWATI DWI PUTRI, S.Pd", "KIM", "XII F-6", "Senin", "07:00", "07:45", "Ruang 1"],
    ["198001012005011001", "KISWATI DWI PUTRI, S.Pd", "KIM", "XII F-6", "Senin", "07:45", "08:30", "Ruang 1"]
]

for r_idx, row_data in enumerate(sample_data, 2):
    for c_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

wb.save("template_jadwal.xlsx")
