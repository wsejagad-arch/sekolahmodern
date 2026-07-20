import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data Siswa"

# Header
headers = ["nis", "nama", "kelas"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4F81BD")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    
# Column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 15

# Borders for headers and 20 empty rows
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for row in ws.iter_rows(min_row=1, max_row=21, min_col=1, max_col=3):
    for cell in row:
        cell.border = thin_border

wb.save("template_siswa.xlsx")
